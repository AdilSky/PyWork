#-*-coding:utf-8-*-
# Time:2017/10/30 23:08
# Author:YangYangJun


import xlrd
import os
from openpyxl.reader.excel import load_workbook
from Wm_UIDemo import readConfig

rC = readConfig.ReadConfig()

class Common(object):
    '公共操作类'

    def __init__(self):
        '初始化函数'
        # 初始化 根目录
        self.rootPath = readConfig.rootPath
        self.dataPath = os.path.join(self.rootPath,'caseData')

    def readExcel(self,**kwargs):
        '读取excel,参数类型为可变关键字参数，字典类型'

        # 提取参数并赋值
        xlsName = kwargs['xlsName']
        sheetName = kwargs['sheetName']
        upList = kwargs['upList']

        self.xlsPath = os.path.join(self.dataPath,xlsName)
        # 打开指定excel
        self.Rb = xlrd.open_workbook(self.xlsPath)
        # 获取 指定sheet
        self.Rs = self.Rb.sheet_by_name(sheetName)
        # 获取表最大行数
        maxRow = self.Rs.nrows
        print maxRow
        # 定义数据字典用于存放需要的数据
        directDataDic = {}
        unDirectDataDic = {}
        for r in range(0,maxRow):
            # 逐行获取指定单元格的值，xlrd 获取单元格是从(0,0)开始
            A = self.Rs.cell(r,upList[0]).value
            H = self.Rs.cell(r,upList[1]).value
            I = self.Rs.cell(r,upList[2]).value
            # 根据指定的值 允许自行，直连用户
            if A != 'CaseNO' and H == 'Y' and I == 'Y':
                # 获取实际行号,然后作为数据字典的key
                row = r + 1
                # 获取整行值作为数据字典的 value
                directDataDic[row] = self.Rs.row_values(r)

            elif A != 'CaseNO' and H == 'Y' and I == 'N':
                # 获取实际行号,然后作为数据字典的key
                row = r + 1
                # 获取整行值作为数据字典的 value
                unDirectDataDic[row] = self.Rs.row_values(r)
        # 返回可用的数据字典
        return directDataDic,unDirectDataDic

    def writeExcel(self,**kwargs):
        '写入excel,参数类型为可变关键字参数，字典类型'

        # 提取参数并赋值
        xlsName = kwargs['xlsName']
        sheetName = kwargs['sheetName']
        dataDic = kwargs['dataDic']
        upList = kwargs['upList']
        # 获取excel的局对路径
        self.xlsPath = os.path.join(self.dataPath,xlsName)
        # 使用openpyxl 的load_workbook 打开excel
        self.Rb = load_workbook(self.xlsPath)
        # 使用get_sheet_by_name 获取指定sheet页
        self.Rs = self.Rb.get_sheet_by_name(sheetName)
        # 遍历dataDic 数据字典的 keys列表
        for r in dataDic.keys():
            # 将数据字典中指定的 数据赋值给 指定的单元格，注意openpyxl 操作单元格是从（1,1）开始的所以对于字典的下标要+1
            self.Rs.cell(row=r,column=upList[0]+1,value=dataDic[r][upList[0]])
            self.Rs.cell(row=r, column=upList[1]+1, value=dataDic[r][upList[1]])
            # 保存 excel,这里使用openpyxl是因为其支持xlsx格式文件的读写。而xlwt不支持，虽然可以写，但是打开后回报格式错误。
            self.Rb.save(self.xlsPath)


if __name__ == '__main__':

    common = Common()
    dataDic = common.readExcel(xlsName='loginCase.xlsx',sheetName='Login',upList=[0,7])
    common.writeExcel(xlsName='loginCase.xlsx', sheetName='Login', dataDic=dataDic, upList=[5, 6])

